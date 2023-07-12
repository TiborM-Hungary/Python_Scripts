import pandas
import openpyxl
from fpdf import FPDF

# Use Pandas to read in the data from the Excel
df = pandas.read_excel('data.xlsx')

# NOTE: be aware of Pandas objects
# You can access rows by index, or columns and their data -> similar fashion like a dictionary
# Print out to see data
# print(df)
# print('##########################################################')

# NOTE: 2 nested loops

# Outer loop: looping rows,
# getting name column/field from the given row -> row['name']

# Inner loop: looping columns,
# getting the given column's title: column.title()
# getting the given row column value for the outer loop: row[column]


# Iterate over data
for index, row in df.iterrows():

    # Create PDF
    pdf = FPDF(orientation='P', unit='pt', format='A4')

    # Add a page
    # Note: if you have multiple pages, then call add_page() for each of them
    pdf.add_page()

    # Each element on the PDF page is a cell, create a cell for each

    # Adding title
    pdf.set_font(family='Times', style='B', size=24)
    pdf.cell(w=0, h=50, txt=row['name'], align='C', border=1, ln=1)

    # Using for loop to generate columns automatically
    for column in df.columns[1:]:
        # Adding another "cell"
        pdf.set_font(family='Times', style='B', size=14)
        pdf.cell(w=100, h=25, txt=f"{column.title()}: ")

        # Adding another "cell"
        pdf.set_font(family='Times', size=14)
        pdf.cell(w=100, h=25, txt=row[column], ln=1)

    # Print output
    pdf.output(f"{row['name']}.pdf")
